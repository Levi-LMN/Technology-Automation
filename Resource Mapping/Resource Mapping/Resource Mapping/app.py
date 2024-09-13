from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_mail import Mail, Message
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime, timedelta
import calendar
import holidays
from sqlalchemy import Boolean  # Import Boolean type if not already imported


app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///data.db'
app.config['SECRET_KEY'] = 'youryguguw827eygsu9oiakio2he'  # Replace this with a strong random key
app.config['MAIL_SERVER'] = 'smtp.gmail.com'  # Replace with your SMTP server
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'technology.automationke@gmail.com'  # Replace with your email
app.config['MAIL_PASSWORD'] = 'zpel wjow murm bhjh'  # Replace with your email password
app.config['MAIL_DEFAULT_SENDER'] = 'Tech Automation <technology.automationke@gmail.com>'  # Replace with your email

db = SQLAlchemy(app)
mail = Mail(app)


# Define Models

class Staff(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    leave_days_remaining = db.Column(db.Float, default=0.0)  # Leave days remaining in hours
    is_team_leader = db.Column(Boolean, default=False)  # New field to indicate if the staff is a team leader
    receive_notifications = db.Column(Boolean, default=True)  # New field for notification preferences
    engagements = db.relationship('Engagement', backref='team_leader', lazy=True)
    proposals = db.relationship('Proposal', backref='team_leader', lazy=True)
    hours_logs = db.relationship('HoursLog', backref='staff_member', lazy=True)
    leave_records = db.relationship('LeaveRecord', backref='staff', lazy=True)
    utilizations = db.relationship('Utilization', backref='staff', lazy=True)


class Engagement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    team_leader_id = db.Column(db.Integer, db.ForeignKey('staff.id'))
    status = db.Column(db.String(10), nullable=False)

class Proposal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    team_leader_id = db.Column(db.Integer, db.ForeignKey('staff.id'))
    status = db.Column(db.String(10), nullable=False)

class NonBillable(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)

class HoursLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    staff_id = db.Column(db.Integer, db.ForeignKey('staff.id'))
    category = db.Column(db.String(20), nullable=False)
    item_id = db.Column(db.Integer, nullable=False)
    hours = db.Column(db.Float, nullable=False)
    date = db.Column(db.Date, nullable=False)  # Changed from week_start to date

class LeaveRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    staff_id = db.Column(db.Integer, db.ForeignKey('staff.id'))
    date = db.Column(db.Date, nullable=False)

    __table_args__ = (db.UniqueConstraint('staff_id', 'date', name='_staff_date_uc'),)

    def __repr__(self):
        return f"LeaveRecord('{self.date}')"

class Utilization(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    staff_id = db.Column(db.Integer, db.ForeignKey('staff.id'))
    week_start = db.Column(db.Date, nullable=False)  # Start of the week
    client_utilization_year_to_date = db.Column(db.Float, default=0.0)
    client_utilization_month_to_date = db.Column(db.Float, default=0.0)
    resource_utilization_year_to_date = db.Column(db.Float, default=0.0)
    resource_utilization_month_to_date = db.Column(db.Float, default=0.0)

    def __repr__(self):
        return (f"Utilization(staff_id={self.staff_id}, week_start={self.week_start}, "
                f"client_utilization_year_to_date={self.client_utilization_year_to_date}, "
                f"client_utilization_month_to_date={self.client_utilization_month_to_date}, "
                f"resource_utilization_year_to_date={self.resource_utilization_year_to_date}, "
                f"resource_utilization_month_to_date={self.resource_utilization_month_to_date})")




@app.route('/send_reminder/<int:member_id>', methods=['POST'])
def send_reminder(member_id):
    member = Staff.query.get_or_404(member_id)

    # Check if the member wants to receive notifications and is not a team leader
    if not member.receive_notifications:
        flash(f"Cannot send reminder to {member.name}: User has opted out of notifications.", "warning")
        return redirect(url_for('view_utilization'))

    if member.is_team_leader:
        flash(f"Cannot send reminder to {member.name}: User is a team leader.", "warning")
        return redirect(url_for('view_utilization'))

    # Get the current week dates
    today = datetime.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)

    # Check if the member has logged utilization and hours for the current week
    last_utilization = Utilization.query.filter_by(staff_id=member.id).order_by(Utilization.week_start.desc()).first()
    last_hours_log = HoursLog.query.filter_by(staff_id=member.id).order_by(HoursLog.date.desc()).first()

    utilization_in_week = last_utilization and start_of_week <= last_utilization.week_start <= end_of_week
    hours_log_in_week = last_hours_log and start_of_week <= last_hours_log.date <= end_of_week

    # Prepare the message based on what the user has or hasn't logged
    if not utilization_in_week and not hours_log_in_week:
        missing_info = "both your utilization numbers and hours log"
    elif not utilization_in_week:
        missing_info = "your utilization numbers"
    elif not hours_log_in_week:
        missing_info = "your hours log"
    else:
        # If both are logged, no reminder needed
        flash(f"No reminder needed for {member.name}: All information is up to date.", "info")
        return redirect(url_for('view_utilization'))

    # URL to the "choose option" page with the staff member's ID
    choose_option_url = url_for('choose_option', staff_id=member.id, _external=True)

    # Styled HTML email content
    msg = Message(
        "Reminder: Please Update Your Weekly Numbers",
        recipients=[member.email]
    )
    msg.html = f"""
    <html>
    <body style="font-family: Arial, sans-serif; background-color: #f4f4f4; padding: 20px;">
        <div style="max-width: 600px; margin: 0 auto; background-color: #ffffff; padding: 20px; border-radius: 8px; box-shadow: 0px 0px 10px rgba(0, 0, 0, 0.1);">
            <h2 style="color: #e3721c; text-align: center;">Reminder to Update Your Weekly Logs</h2>
            <p style="font-size: 16px; color: #333;">
                Dear <strong>{member.name}</strong>,
            </p>
            <p style="font-size: 16px; color: #333;">
                We noticed that you have not yet updated <strong>{missing_info}</strong> for the current week 
                ({start_of_week.strftime('%B %d')} to {end_of_week.strftime('%B %d')}).
            </p>
            <p style="font-size: 16px; color: #333;">
                Please log in to the system at your earliest convenience to make the necessary updates.
            </p>
            <div style="text-align: center; margin: 20px 0;">
                <a href="{choose_option_url}" style="background-color: #e3721c; color: #ffffff; padding: 12px 20px; border-radius: 5px; text-decoration: none; font-size: 16px;">
                    Update Now
                </a>
            </div>
            <p style="font-size: 12px; color: #999; text-align: center;">
                This is an automated email. Please do not reply to this message.
            </p>
        </div>
    </body>
    </html>
    """

    try:
        mail.send(msg)
        flash(f"Reminder sent to {member.name} for missing {missing_info}.", "success")
    except Exception as e:
        flash(f"Failed to send reminder to {member.name}: {str(e)}", "error")

    return redirect(url_for('view_utilization'))

@app.route('/send_all_reminders', methods=['POST'])
def send_all_reminders():
    members = Staff.query.all()
    reminders_sent = 0
    for member in members:
        # Check if the member wants to receive notifications and is not a team leader
        if not member.receive_notifications:
            continue
        if member.is_team_leader:
            continue

        # Check if the member has logged utilization and hours for the current week
        today = datetime.now().date()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)

        last_utilization = Utilization.query.filter_by(staff_id=member.id).order_by(
            Utilization.week_start.desc()).first()
        last_hours_log = HoursLog.query.filter_by(staff_id=member.id).order_by(HoursLog.date.desc()).first()

        utilization_in_week = last_utilization and start_of_week <= last_utilization.week_start <= end_of_week
        hours_log_in_week = last_hours_log and start_of_week <= last_hours_log.date <= end_of_week

        # Prepare the message based on what the user has or hasn't logged
        if not utilization_in_week or not hours_log_in_week:
            missing_info = ""
            if not utilization_in_week and not hours_log_in_week:
                missing_info = "both your utilization numbers and hours log"
            elif not utilization_in_week:
                missing_info = "your utilization numbers"
            elif not hours_log_in_week:
                missing_info = "your hours log"

            # URL to the "choose option" page with the staff member's ID
            choose_option_url = url_for('choose_option', staff_id=member.id, _external=True)

            # Prepare and send the email
            msg = Message(
                "Reminder: Please Update Your Weekly Numbers",
                recipients=[member.email]
            )
            msg.html = f"""
            <html>
            <body style="font-family: Arial, sans-serif; background-color: #f4f4f4; padding: 20px;">
                <div style="max-width: 600px; margin: 0 auto; background-color: #ffffff; padding: 20px; border-radius: 8px; box-shadow: 0px 0px 10px rgba(0, 0, 0, 0.1);">
                    <h2 style="color: #e3721c; text-align: center;">Reminder to Update Your Weekly Logs</h2>
                    <p style="font-size: 16px; color: #333;">
                        Dear <strong>{member.name}</strong>,
                    </p>
                    <p style="font-size: 16px; color: #333;">
                        We noticed that you have not yet updated <strong>{missing_info}</strong> for the current week 
                        ({start_of_week.strftime('%B %d')} to {end_of_week.strftime('%B %d')}).
                    </p>
                    <p style="font-size: 16px; color: #333;">
                        Please log in to the system at your earliest convenience to make the necessary updates.
                    </p>
                    <div style="text-align: center; margin: 20px 0;">
                        <a href="{choose_option_url}" style="background-color: #e3721c; color: #ffffff; padding: 12px 20px; border-radius: 5px; text-decoration: none; font-size: 16px;">
                            Update Now
                        </a>
                    </div>
                    <p style="font-size: 12px; color: #999; text-align: center;">
                        This is an automated email. Please do not reply to this message.
                    </p>
                </div>
            </body>
            </html>
            """

            try:
                mail.send(msg)
                reminders_sent += 1
            except Exception as e:
                flash(f"Failed to send reminder to {member.name}: {str(e)}", "error")

    flash(f"Reminders sent to {reminders_sent} eligible members.", "success")
    return redirect(url_for('view_utilization'))

@app.route('/members')
def members():
    today = datetime.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)

    members = Staff.query.all()
    member_data = []

    for member in members:
        last_utilization = Utilization.query.filter_by(staff_id=member.id).order_by(Utilization.week_start.desc()).first()
        last_hours_log = HoursLog.query.filter_by(staff_id=member.id).order_by(HoursLog.date.desc()).first()

        utilization_in_week = last_utilization and start_of_week <= last_utilization.week_start <= end_of_week
        hours_log_in_week = last_hours_log and start_of_week <= last_hours_log.date <= end_of_week

        member_data.append({
            'id': member.id,
            'name': member.name,
            'email': member.email,
            'last_utilization_week': last_utilization.week_start if last_utilization else None,
            'last_hours_log_date': last_hours_log.date if last_hours_log else None,
            'show_reminder': not (utilization_in_week and hours_log_in_week)
        })

    return render_template('members.html', members=member_data)



@app.route('/choose_option/<int:staff_id>')
def choose_option(staff_id):
    staff = Staff.query.get_or_404(staff_id)
    return render_template('choose_option.html', staff=staff)


@app.route('/staff_details/<int:staff_id>', methods=['GET', 'POST'])
def staff_details(staff_id):
    staff_member = Staff.query.get(staff_id)

    if not staff_member:
        flash('Staff member not found.', 'error')
        return redirect(url_for('edit_data'))  # Redirect if not found

    if request.method == 'POST':
        # Handle the toggle action if a form is submitted
        if 'toggle_notifications' in request.form:
            staff_member.receive_notifications = not staff_member.receive_notifications
            db.session.commit()
            flash('Notification preferences updated successfully.', 'success')

    return render_template('staff_details.html', staff_member=staff_member)



@app.route('/update_utilization/<int:staff_id>', methods=['GET', 'POST'])
def update_utilization(staff_id):
    staff = Staff.query.get_or_404(staff_id)
    
    if request.method == 'POST':
        if 'update_utilization' in request.form:
            week_start = datetime.strptime(request.form['week_start'], '%Y-%m-%d').date()
            
            # Update or create Utilization record
            utilization = Utilization.query.filter_by(staff_id=staff_id, week_start=week_start).first()
            if not utilization:
                utilization = Utilization(staff_id=staff_id, week_start=week_start)
            
            utilization.client_utilization_year_to_date = float(request.form['client_utilization_year_to_date'])
            utilization.client_utilization_month_to_date = float(request.form['client_utilization_month_to_date'])
            utilization.resource_utilization_year_to_date = float(request.form['resource_utilization_year_to_date'])
            utilization.resource_utilization_month_to_date = float(request.form['resource_utilization_month_to_date'])
            
            db.session.add(utilization)
            db.session.commit()
            flash('Utilization updated successfully!', 'success')
        
        elif 'update_leave_days' in request.form:
            # Update staff leave days remaining
            old_leave_days = staff.leave_days_remaining
            new_leave_days = float(request.form['leave_days_remaining'])
            
            # Calculate the difference in days
            days_difference = (old_leave_days - new_leave_days) / 7.5  # Assuming 7.5 hours per day
            
            if days_difference != 0:
                if days_difference > 0:
                    # Add new leave records (decrease in available leave)
                    for _ in range(int(days_difference)):
                        new_leave_date = find_next_workday(staff_id)
                        new_leave = LeaveRecord(staff_id=staff_id, date=new_leave_date)
                        try:
                            db.session.add(new_leave)
                            db.session.flush()
                        except IntegrityError:
                            db.session.rollback()
                            flash(f"Couldn't add leave for {new_leave_date}. Date already exists.", 'warning')
                else:
                    # Remove leave records (increase in available leave)
                    leave_records = LeaveRecord.query.filter_by(staff_id=staff_id).order_by(LeaveRecord.date.desc()).limit(int(abs(days_difference))).all()
                    for record in leave_records:
                        db.session.delete(record)
            
            staff.leave_days_remaining = new_leave_days
            
            try:
                db.session.commit()
                flash('Leave days updated successfully!', 'success')
            except IntegrityError:
                db.session.rollback()
                flash('Error updating leave records. Please try again.', 'error')
        
        return redirect(url_for('update_utilization', staff_id=staff_id))

    week_start = datetime.now().date() - timedelta(days=datetime.now().weekday())
    week_end = week_start + timedelta(days=4)
    
    # Get the latest utilization record
    latest_utilization = Utilization.query.filter_by(staff_id=staff_id).order_by(Utilization.week_start.desc()).first()
    
    return render_template('update_utilization.html', staff=staff, week_start=week_start, week_end=week_end, latest_utilization=latest_utilization)


@app.route('/manage_leave/<int:staff_id>', methods=['GET', 'POST'])
def manage_leave(staff_id):
    staff_member = Staff.query.get(staff_id)
    leave_records = LeaveRecord.query.filter_by(staff_id=staff_id).all()

    if request.method == 'POST':
        leave_date = request.form['leave_date']
        existing_record = LeaveRecord.query.filter_by(staff_id=staff_id, date=leave_date).first()
        if existing_record:
            flash('You have already taken leave on this day.', 'error')
        else:
            new_record = LeaveRecord(staff_id=staff_id, date=leave_date)
            db.session.add(new_record)
            db.session.commit()
            flash('Leave record added successfully.', 'success')
        return redirect(url_for('manage_leave', staff_id=staff_id))

    return render_template('manage_leave.html', staff_member=staff_member, leave_records=leave_records)




from sqlalchemy.exc import IntegrityError




def find_next_workday(staff_id):
    last_leave = LeaveRecord.query.filter_by(staff_id=staff_id).order_by(LeaveRecord.date.desc()).first()
    if last_leave:
        start_date = last_leave.date + timedelta(days=1)
    else:
        start_date = datetime.now().date()
    
    while True:
        if start_date.weekday() < 5:  # Monday = 0, Friday = 4
            return start_date
        start_date += timedelta(days=1)

@app.route('/get_staff_data/<int:staff_id>/<string:week_start>')
def get_staff_data(staff_id, week_start):
    staff = Staff.query.get_or_404(staff_id)
    week_start_date = datetime.strptime(week_start, '%Y-%m-%d').date()
    utilization = Utilization.query.filter_by(staff_id=staff_id, week_start=week_start_date).first()
    
    data = {
        'leave_days_remaining': staff.leave_days_remaining
    }
    
    if utilization:
        data.update({
            'client_utilization_year_to_date': utilization.client_utilization_year_to_date,
            'client_utilization_month_to_date': utilization.client_utilization_month_to_date,
            'resource_utilization_year_to_date': utilization.resource_utilization_year_to_date,
            'resource_utilization_month_to_date': utilization.resource_utilization_month_to_date
        })
    else:
        data.update({
            'client_utilization_year_to_date': '',
            'client_utilization_month_to_date': '',
            'resource_utilization_year_to_date': '',
            'resource_utilization_month_to_date': ''
        })
    
    return jsonify(data)

from datetime import date

@app.route('/view_utilization')
def view_utilization():
    # Get current week's start (assuming week starts on Monday)
    current_week_start = date.today() - timedelta(days=date.today().weekday())

    # Get the most recent utilization record per staff member
    utilizations = (db.session.query(Utilization)
                    .group_by(Utilization.staff_id)
                    .order_by(Utilization.staff_id, Utilization.week_start.desc())
                    .all())

    return render_template('view_utilization.html', utilizations=utilizations, current_week_start=current_week_start)


@app.route('/')
def index():
    return render_template('index.html')

from collections import defaultdict
from datetime import datetime, timedelta
from flask import render_template

@app.route('/view_details')
def view_details():
    # Query for all necessary data
    staff_members = Staff.query.options(
        db.joinedload(Staff.engagements),
        db.joinedload(Staff.proposals),
        db.joinedload(Staff.hours_logs)
    ).all()
    engagements = Engagement.query.all()
    proposals = Proposal.query.all()
    non_billables = NonBillable.query.all()
    logs = HoursLog.query.all()

    # Create a dictionary mapping staff IDs to names
    staff_names = {staff.id: staff.name for staff in staff_members}

    # Process logs to summarize by staff and week
    weekly_summary = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    for log in logs:
        staff_name = staff_names.get(log.staff_id, 'Unknown')
        week_start = log.date - timedelta(days=log.date.weekday())
        daily_summary = weekly_summary[staff_name][week_start]
        daily_summary[log.date] += log.hours

    # Convert weekly_summary to a list of dictionaries for easier template rendering
    summary_list = []
    for staff, weeks in weekly_summary.items():
        for week_start, days in weeks.items():
            week_summary = {
                'staff': staff,
                'week_start': week_start,
                'daily_summary': sorted(days.items())  # Sort by date
            }
            summary_list.append(week_summary)

    # Pass all data to the template
    return render_template('view_details.html',
                           staff_members=staff_members,
                           engagements=engagements,
                           proposals=proposals,
                           non_billables=non_billables,
                           logs=logs,
                           staff_names=staff_names,
                           weekly_summary=summary_list)



@app.route('/add_staff', methods=['GET', 'POST'])
def add_staff():
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')  # New line for email
        staff = Staff(name=name, email=email)  # Update instantiation
        db.session.add(staff)
        db.session.commit()
        return redirect(url_for('index'))
    return render_template('add_staff.html')


@app.route('/add_engagement', methods=['GET', 'POST'])
def add_engagement():
    if request.method == 'POST':
        name = request.form.get('name')
        team_leader_id = request.form.get('team_leader_id')
        status = request.form.get('status')
        
        # Create new engagement
        engagement = Engagement(name=name, team_leader_id=team_leader_id, status=status)
        db.session.add(engagement)
        
        # Update is_team_leader for the selected staff member
        if team_leader_id:
            staff_member = Staff.query.get(team_leader_id)
            if staff_member:
                staff_member.is_team_leader = True  # Set to True if selected as team leader
                db.session.add(staff_member)
        
        db.session.commit()
        return redirect(url_for('index'))
    
    staff_members = Staff.query.all()
    return render_template('add_engagement.html', staff_members=staff_members)

@app.route('/add_proposal', methods=['GET', 'POST'])
def add_proposal():
    if request.method == 'POST':
        name = request.form.get('name')
        team_leader_id = request.form.get('team_leader_id')
        status = request.form.get('status')
        
        # Create new proposal
        proposal = Proposal(name=name, team_leader_id=team_leader_id, status=status)
        db.session.add(proposal)
        
        # Update is_team_leader for the selected staff member
        if team_leader_id:
            staff_member = Staff.query.get(team_leader_id)
            if staff_member:
                staff_member.is_team_leader = True  # Set to True if selected as team leader
                db.session.add(staff_member)
        
        db.session.commit()
        return redirect(url_for('index'))
    
    staff_members = Staff.query.all()
    return render_template('add_proposal.html', staff_members=staff_members)


@app.route('/add_non_billable', methods=['GET', 'POST'])
def add_non_billable():
    if request.method == 'POST':
        name = request.form.get('name')
        non_billable = NonBillable(name=name)
        db.session.add(non_billable)
        db.session.commit()
        return redirect(url_for('index'))
    return render_template('add_non_billable.html')

@app.route('/log_hours/<int:staff_member_id>', methods=['GET', 'POST'])
def log_hours(staff_member_id):
    staff_member = Staff.query.get_or_404(staff_member_id)
    proposals = Proposal.query.filter_by(status='Active').all()
    engagements = Engagement.query.filter_by(status='Active').all()
    non_billables = NonBillable.query.all()

    if request.method == 'POST':
        try:
            week_start = datetime.strptime(request.form.get('week_start'), '%Y-%m-%d').date()
        except ValueError:
            flash('Invalid week start date format. Please use YYYY-MM-DD.', 'error')
            return redirect(url_for('log_hours', staff_member_id=staff_member_id))

        errors = []

        categories = [
            ('proposal', 'proposal_ids', 'proposal_hours'),
            ('engagement', 'engagement_ids', 'engagement_hours'),
            ('non_billable', 'nonbillable_ids', 'nonbillable_hours')
        ]

        hours_logged = False
        for day in range(5):  # Monday to Friday
            current_date = week_start + timedelta(days=day)

            # Delete existing logs for this date
            try:
                HoursLog.query.filter_by(
                    staff_id=staff_member.id,
                    date=current_date
                ).delete()
            except Exception as e:
                flash(f'Error deleting existing logs for {current_date}: {str(e)}', 'error')
                return redirect(url_for('log_hours', staff_member_id=staff_member_id))

            for category, id_field, hours_field in categories:
                ids = request.form.getlist(f'{id_field}[]')
                hours = request.form.to_dict()

                for item_id in ids:
                    item_hours = hours.get(f"{hours_field}[{item_id}][{day}]")
                    if item_hours:
                        try:
                            item_hours = float(item_hours)
                            hours_log = HoursLog(
                                staff_id=staff_member.id,
                                category=category,
                                item_id=int(item_id),
                                hours=item_hours,
                                date=current_date
                            )
                            db.session.add(hours_log)
                            hours_logged = True
                        except ValueError:
                            errors.append(f"Invalid hours format for {category} item {item_id} on {current_date}.")
                        except Exception as e:
                            flash(f'Error logging hours for {category} item {item_id} on {current_date}: {str(e)}', 'error')
                            return redirect(url_for('log_hours', staff_member_id=staff_member_id))

        if errors:
            for error in errors:
                flash(error, 'error')
            return redirect(url_for('log_hours', staff_member_id=staff_member_id))

        if not hours_logged:
            flash('Please log hours for at least one day and category.', 'error')
            return redirect(url_for('log_hours', staff_member_id=staff_member_id))

        try:
            db.session.commit()
        except Exception as e:
            flash(f'Error committing changes: {str(e)}', 'error')
            return redirect(url_for('log_hours', staff_member_id=staff_member_id))

        flash('Hours successfully logged for the week.', 'success')
        return redirect(url_for('thank_you', staff_id=staff_member_id))


    # For GET requests
    week_start = request.args.get('week_start')
    if week_start:
        try:
            week_start = datetime.strptime(week_start, '%Y-%m-%d').date()
        except ValueError:
            flash('Invalid week start date format. Please use YYYY-MM-DD.', 'error')
            return redirect(url_for('log_hours', staff_member_id=staff_member_id))
    else:
        today = datetime.now().date()
        week_start = today - timedelta(days=today.weekday())

    # Retrieve existing logs for the selected week
    try:
        existing_logs = HoursLog.query.filter(
            HoursLog.staff_id == staff_member.id,
            HoursLog.date >= week_start,
            HoursLog.date < week_start + timedelta(days=5)
        ).all()
    except Exception as e:
        flash(f'Error retrieving existing logs: {str(e)}', 'error')
        return redirect(url_for('log_hours', staff_member_id=staff_member_id))

    existing_hours = {}
    for log in existing_logs:
        day_index = (log.date - week_start).days
        if 0 <= day_index < 5:  # Ensure it's within the work week
            # Change 'non_billable' to 'nonbillable' to match the HTML
            category = 'nonbillable' if log.category == 'non_billable' else log.category
            existing_hours.setdefault(category, {}).setdefault(log.item_id, {})[day_index] = log.hours

    # Add debug print statements
    print("Existing hours:", existing_hours)
    print("Non-billables:", non_billables)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # If it's an AJAX request, return JSON data
        return jsonify({
            'existing_hours': existing_hours,
            'week_start': week_start.strftime('%Y-%m-%d')
        })

    return render_template(
        'log_hours.html',
        staff_member=staff_member,
        proposals=proposals,
        engagements=engagements,
        non_billables=non_billables,
        existing_hours=existing_hours,
        week_start=week_start.strftime('%Y-%m-%d')
    )

@app.route('/generate_excel')
def generate_excel():
    wb = Workbook()

    # Color codes and styles
    proposal_color = PatternFill(start_color="FFB3E5FC", end_color="FFB3E5FC", fill_type="solid")
    engagement_color = PatternFill(start_color="FFFFCCCB", end_color="FFFFCCCB", fill_type="solid")
    non_billable_color = PatternFill(start_color="FFD7F6D3", end_color="FFD7F6D3", fill_type="solid")
    total_color = PatternFill(start_color="FFFFD700", end_color="FFFFD700", fill_type="solid")
    header_fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
    leave_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    header_font = Font(bold=True, size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def style_sheet(ws, color_fill):
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                else:
                    cell.fill = color_fill

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Proposals sheet
    ws_proposals = wb.active
    ws_proposals.title = "Proposals"
    ws_proposals.append(["ID", "Name", "Team Leader", "Status"])
    for proposal in Proposal.query.all():
        ws_proposals.append([proposal.id, proposal.name, proposal.team_leader.name if proposal.team_leader else "N/A", proposal.status])
    style_sheet(ws_proposals, proposal_color)

    # Engagements sheet
    ws_engagements = wb.create_sheet("Engagements")
    ws_engagements.append(["ID", "Name", "Team Leader", "Status"])
    for engagement in Engagement.query.all():
        ws_engagements.append([engagement.id, engagement.name, engagement.team_leader.name if engagement.team_leader else "N/A", engagement.status])
    style_sheet(ws_engagements, engagement_color)

    # Non-Billables sheet
    ws_non_billables = wb.create_sheet("Non-Billables")
    ws_non_billables.append(["ID", "Name"])
    for non_billable in NonBillable.query.all():
        ws_non_billables.append([non_billable.id, non_billable.name])
    style_sheet(ws_non_billables, non_billable_color)

    # Staff Members sheet
    ws_staff = wb.create_sheet("Staff Members")

    # Get all unique months from HoursLog
    unique_months = db.session.query(
        db.func.strftime('%Y-%m', HoursLog.date).label('month')
    ).distinct().order_by('month').all()

    current_row = 1

    # Get all staff members
    all_staff = Staff.query.all()

    for month_tuple in unique_months:
        month = datetime.strptime(month_tuple[0], '%Y-%m')
        month_start = month.replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        ws_staff.cell(row=current_row, column=1, value=f"Month: {month.strftime('%B %Y')}")
        ws_staff.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 1

        headers = ["ID", "Name", "Engagements", "Proposals", "Non-Billables", "Total Hours", "Category"]
        for col, header in enumerate(headers, start=1):
            cell = ws_staff.cell(row=current_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        current_row += 1

        for staff in all_staff:
            month_logs = HoursLog.query.filter(
                HoursLog.staff_id == staff.id,
                HoursLog.date >= month_start,
                HoursLog.date <= month_end
            ).all()

            if month_logs:
                logged_engagements = set(
                    Engagement.query.get(log.item_id).name for log in month_logs if log.category == 'engagement')
                logged_proposals = set(
                    Proposal.query.get(log.item_id).name for log in month_logs if log.category == 'proposal')

                engagements = ", ".join(logged_engagements)
                proposals = ", ".join(logged_proposals)
                non_billable_hours = sum([log.hours for log in month_logs if log.category == 'non_billable'])
                total_hours = sum([log.hours for log in month_logs])

                if logged_engagements:
                    primary_category = "Engagement"
                elif logged_proposals:
                    primary_category = "Proposal"
                else:
                    primary_category = "Non-Billable"
            else:
                engagements = ""
                proposals = ""
                non_billable_hours = 0
                total_hours = 0
                primary_category = ""

            row_data = [
                staff.id,
                staff.name,
                engagements,
                proposals,
                f"{non_billable_hours:.2f} hours" if non_billable_hours > 0 else "",
                f"{total_hours:.2f} hours" if total_hours > 0 else "",
                primary_category
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws_staff.cell(row=current_row, column=col, value=value)
                if col == 7:  # Category column
                    if value == "Engagement":
                        cell.fill = engagement_color
                    elif value == "Proposal":
                        cell.fill = proposal_color
                    elif value == "Non-Billable":
                        cell.fill = non_billable_color

            current_row += 1

        current_row += 2  # Add space between tables

    # Adjust column widths
    for column in ws_staff.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_staff.column_dimensions[column_letter].width = adjusted_width

    # Generate monthly sheets
    ke_holidays = holidays.KE()

    earliest_date = db.session.query(db.func.min(HoursLog.date)).scalar()
    if earliest_date is None:
        earliest_date = datetime.now().date().replace(day=1)

    current_year = datetime.now().year
    if datetime.now().month > 6:
        end_financial_year = datetime(current_year + 1, 6, 30).date()
    else:
        end_financial_year = datetime(current_year, 6, 30).date()

    # ... (previous code remains the same)

    current_date = earliest_date.replace(day=1)
    while current_date <= end_financial_year:
        month_name = current_date.strftime('%B %Y')
        ws_month = wb.create_sheet(month_name)

        header_fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
        subheader_fill = PatternFill(start_color="FFF0F0F0", end_color="FFF0F0F0", fill_type="solid")
        holiday_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True, size=12)
        subheader_font = Font(bold=True, size=11)

        columns = ["Staff Name"]
        categories = ["Proposals", "Engagements", "Non-Billables", "Total"]

        working_days = []
        month_end = current_date.replace(day=calendar.monthrange(current_date.year, current_date.month)[1])
        current = max(current_date, earliest_date)
        while current <= month_end:
            if current.weekday() < 5:
                working_days.append(current)
            current += timedelta(days=1)

        for day in working_days:
            date_header = f"{day.strftime('%A %d %B %Y')}"
            columns.extend([date_header] + ['' for _ in range(len(categories) - 1)])
        ws_month.append(columns)

        sub_headers = [""]
        for _ in range(len(working_days)):
            sub_headers.extend(categories)
        ws_month.append(sub_headers)

        # Style and format headers
        for row in ws_month[1:3]:
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col in range(1, len(columns) + 1):
            header_cell = ws_month.cell(row=1, column=col)
            subheader_cell = ws_month.cell(row=2, column=col)

            header_cell.fill = header_fill
            header_cell.font = header_font
            subheader_cell.fill = subheader_fill
            subheader_cell.font = subheader_font

            if col > 1 and (col - 2) % 4 == 0:
                ws_month.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)

            if subheader_cell.value == "Proposals":
                subheader_cell.fill = proposal_color
            elif subheader_cell.value == "Engagements":
                subheader_cell.fill = engagement_color
            elif subheader_cell.value == "Non-Billables":
                subheader_cell.fill = non_billable_color
            elif subheader_cell.value == "Total":
                subheader_cell.fill = total_color

            if col > 1 and (col - 2) % 4 == 0:
                date_str = header_cell.value.split(' ', 1)[1]
                date = datetime.strptime(date_str, '%d %B %Y').date()
                if date in ke_holidays:
                    header_cell.fill = holiday_fill
                    header_cell.font = Font(bold=True, color="FFFFFFFF")
                    ws_month.cell(row=1, column=col, value=f"{header_cell.value} (Holiday: {ke_holidays.get(date)})")

        ws_month.row_dimensions[1].height = 30
        ws_month.row_dimensions[2].height = 25

        # Start adding data from row 3
        for idx, staff in enumerate(Staff.query.all(), start=3):
            row_data = [staff.name]

            for day in working_days:
                leave_record = LeaveRecord.query.filter_by(staff_id=staff.id, date=day).first()
                if leave_record:
                    row_data.extend(["ON LEAVE", "", "", ""])
                else:
                    logs = HoursLog.query.filter_by(staff_id=staff.id, date=day).all()
                    daily_total = 0
                    for category in ['proposal', 'engagement', 'non_billable']:
                        category_logs = [log for log in logs if log.category == category]
                        category_hours = sum([log.hours for log in category_logs])
                        if category_hours > 0:
                            items = []
                            for log in category_logs:
                                if category == 'proposal':
                                    item = Proposal.query.get(log.item_id)
                                elif category == 'engagement':
                                    item = Engagement.query.get(log.item_id)
                                else:
                                    item = NonBillable.query.get(log.item_id)
                                items.append(f"{item.name} ({log.hours:.2f}h)")
                            row_data.append("\n".join(items))
                            daily_total += category_hours
                        else:
                            row_data.append("")
                    row_data.append(f"{daily_total:.2f}" if daily_total > 0 else "")

            # Add the row data to the sheet
            for col, value in enumerate(row_data, start=1):
                cell = ws_month.cell(row=idx, column=col, value=value)
                if value == "ON LEAVE":
                    for i in range(4):
                        leave_cell = ws_month.cell(row=idx, column=col + i)
                        leave_cell.fill = leave_fill
                        if i == 0:
                            leave_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Apply borders and alignment to all cells
        for row in ws_month[3:ws_month.max_row + 1]:
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Adjust column widths
        for i, column in enumerate(ws_month.columns, 1):
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_month.column_dimensions[get_column_letter(i)].width = adjusted_width

        ws_month.freeze_panes = "B3"

        current_date = (current_date + timedelta(days=32)).replace(day=1)



     # Generate the file name based on the content and current date
    today = datetime.now().strftime('%Y-%m-%d')
    file_name = f"Resource Mapping and Leave Days - {today}.xlsx"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name=file_name, as_attachment=True)




@app.route('/thank_you/<int:staff_id>')
def thank_you(staff_id):
    staff = Staff.query.get_or_404(staff_id)
    return render_template('thankyou.html', staff=staff)



from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime, timedelta
import calendar
import holidays

# ... (existing code)

@app.route('/edit_data', methods=['GET', 'POST'])
def edit_data():
    if request.method == 'POST':
        data_type = request.form.get('type')
        item_id = request.form.get('id')

        if 'delete' in request.form:
            delete_value = request.form.get('delete')
            delete_type, delete_id = delete_value.split('-')

            if delete_type == 'staff':
                item = Staff.query.get(delete_id)
            elif delete_type == 'engagement':
                item = Engagement.query.get(delete_id)
            elif delete_type == 'proposal':
                item = Proposal.query.get(delete_id)
            elif delete_type == 'non_billable':
                item = NonBillable.query.get(delete_id)

            if item:
                db.session.delete(item)
                db.session.commit()
                flash(f'{delete_type.capitalize()} deleted successfully.', 'success')
            else:
                flash(f'{delete_type.capitalize()} not found.', 'error')

        else:
            if data_type == 'staff':
                staff = Staff.query.get(item_id)
                if staff:
                    staff.name = request.form.get('name')
                    staff.email = request.form.get('email')  # Update to use email
                    db.session.commit()
                    flash('Staff member updated successfully.', 'success')

            elif data_type == 'engagement':
                engagement = Engagement.query.get(item_id)
                if engagement:
                    old_team_leader_id = engagement.team_leader_id
                    engagement.name = request.form.get('name')
                    engagement.team_leader_id = request.form.get('team_leader_id')
                    engagement.status = request.form.get('status')

                    # Check if old team leader is still leading any engagement or proposal
                    if old_team_leader_id != engagement.team_leader_id:
                        old_team_leader = Staff.query.get(old_team_leader_id)
                        new_team_leader = Staff.query.get(engagement.team_leader_id)

                        # Check if the old team leader is leading any other engagements or proposals
                        if not Engagement.query.filter(Engagement.team_leader_id == old_team_leader_id).filter(Engagement.id != engagement.id).count() and \
                           not Proposal.query.filter(Proposal.team_leader_id == old_team_leader_id).count():
                            if old_team_leader:
                                old_team_leader.is_team_leader = False
                                db.session.add(old_team_leader)

                        # Set is_team_leader for the new team leader
                        if new_team_leader:
                            new_team_leader.is_team_leader = True
                            db.session.add(new_team_leader)

                    db.session.commit()
                    flash('Engagement updated successfully.', 'success')

            elif data_type == 'proposal':
                proposal = Proposal.query.get(item_id)
                if proposal:
                    old_team_leader_id = proposal.team_leader_id
                    proposal.name = request.form.get('name')
                    proposal.team_leader_id = request.form.get('team_leader_id')
                    proposal.status = request.form.get('status')

                    # Check if old team leader is still leading any engagement or proposal
                    if old_team_leader_id != proposal.team_leader_id:
                        old_team_leader = Staff.query.get(old_team_leader_id)
                        new_team_leader = Staff.query.get(proposal.team_leader_id)

                        # Check if the old team leader is leading any other engagements or proposals
                        if not Engagement.query.filter(Engagement.team_leader_id == old_team_leader_id).count() and \
                           not Proposal.query.filter(Proposal.team_leader_id == old_team_leader_id).filter(Proposal.id != proposal.id).count():
                            if old_team_leader:
                                old_team_leader.is_team_leader = False
                                db.session.add(old_team_leader)

                        # Set is_team_leader for the new team leader
                        if new_team_leader:
                            new_team_leader.is_team_leader = True
                            db.session.add(new_team_leader)

                    db.session.commit()
                    flash('Proposal updated successfully.', 'success')

            elif data_type == 'non_billable':
                non_billable = NonBillable.query.get(item_id)
                if non_billable:
                    non_billable.name = request.form.get('name')
                    db.session.commit()
                    flash('Non-Billable updated successfully.', 'success')

        return redirect(url_for('edit_data'))

    staff_members = Staff.query.all()
    engagements = Engagement.query.all()
    proposals = Proposal.query.all()
    non_billables = NonBillable.query.all()

    return render_template('edit_data.html',
                           staff_members=staff_members,
                           engagements=engagements,
                           proposals=proposals,
                           non_billables=non_billables)

# ... (rest of the existing code)

@app.route('/staff_overview', methods=['GET'])
def staff_overview():
    staff_members = Staff.query.all()
    return render_template('staff_overview.html', staff_members=staff_members)



from datetime import datetime, timedelta

@app.route('/leave_balances')
def leave_balances():
    staff_members = Staff.query.all()
    current_date = datetime.now().date()

    # Determine the start of the current financial year
    if current_date.month > 6:
        start_of_financial_year = datetime(current_date.year, 7, 1).date()
    else:
        start_of_financial_year = datetime(current_date.year - 1, 7, 1).date()

    balances = []
    for staff in staff_members:
        # Use the leave_days_remaining from the Staff model
        current_balance_hours = staff.leave_days_remaining
        current_balance_days = current_balance_hours / 7.5  # Convert hours to days

        # Calculate leave taken this financial year
        leave_taken = LeaveRecord.query.filter(
            LeaveRecord.staff_id == staff.id,
            LeaveRecord.date >= start_of_financial_year
        ).count()

        balances.append({
            'staff': staff,
            'used_leave_days': leave_taken,
            'used_leave_hours': leave_taken * 7.5,  # Assuming 7.5 hours per day
            'current_balance_days': round(current_balance_days, 2),
            'current_balance_hours': round(current_balance_hours, 2)
        })

    return render_template('leave_balances.html', balances=balances)

from flask import render_template, request, redirect, url_for, flash

from sqlalchemy.exc import IntegrityError

@app.route('/edit_leave/<int:staff_id>', methods=['GET', 'POST'])
def edit_leave(staff_id):
    staff_member = Staff.query.get_or_404(staff_id)
    leave_records = LeaveRecord.query.filter_by(staff_id=staff_id).order_by(LeaveRecord.date).all()

    if request.method == 'POST':
        if 'delete' in request.form:
            # Delete leave record
            leave_id = int(request.form['delete'])
            leave_record = LeaveRecord.query.get(leave_id)
            if leave_record and leave_record.staff_id == staff_id:
                db.session.delete(leave_record)
                staff_member.leave_days_remaining += 7.5  # Add back 7.5 hours
                db.session.commit()
                flash('Leave record deleted successfully.', 'success')
            else:
                flash('Invalid leave record.', 'error')
        elif 'edit' in request.form:
            # Edit leave record (date change doesn't affect hours)
            leave_id = int(request.form['edit'])
            new_date = request.form.get(f'new_date_{leave_id}')
            leave_record = LeaveRecord.query.get(leave_id)
            if leave_record and leave_record.staff_id == staff_id:
                try:
                    leave_record.date = datetime.strptime(new_date, '%Y-%m-%d').date()
                    db.session.commit()
                    flash('Leave record updated successfully.', 'success')
                except ValueError:
                    flash('Invalid date format. Please use YYYY-MM-DD.', 'error')
                except IntegrityError:
                    db.session.rollback()
                    flash('A leave record for this date already exists.', 'error')
            else:
                flash('Invalid leave record.', 'error')
        elif 'add' in request.form:
            # Add new leave record
            new_leave_date = request.form.get('new_leave_date')
            try:
                new_date = datetime.strptime(new_leave_date, '%Y-%m-%d').date()
                new_leave = LeaveRecord(staff_id=staff_id, date=new_date)
                db.session.add(new_leave)
                staff_member.leave_days_remaining -= 7.5  # Subtract 7.5 hours
                db.session.commit()
                flash('New leave record added successfully.', 'success')
            except ValueError:
                flash('Invalid date format. Please use YYYY-MM-DD.', 'error')
            except IntegrityError:
                db.session.rollback()
                flash('A leave record for this date already exists.', 'error')

        return redirect(url_for('edit_leave', staff_id=staff_id))

    return render_template('edit_leave.html', staff_member=staff_member, leave_records=leave_records)

from collections import defaultdict
from datetime import datetime, timedelta

from collections import defaultdict
from datetime import datetime, timedelta


from datetime import datetime, timedelta

@app.route('/staff_detail/<int:staff_id>')
def staff_detail(staff_id):
    staff = Staff.query.get_or_404(staff_id)

    engagements = Engagement.query.filter_by(team_leader_id=staff_id).all()
    proposals = Proposal.query.filter_by(team_leader_id=staff_id).all()
    hours_logs = HoursLog.query.filter_by(staff_id=staff_id).order_by(HoursLog.date).all()

    cumulative_hours = {
        'engagements': defaultdict(float),
        'proposals': defaultdict(float),
        'non_billables': defaultdict(float)
    }

    hours_logs_by_date = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list))))

    total_hours = 0
    categories_hours = defaultdict(float)
    first_log_date = None
    last_log_date = None

    for log in hours_logs:
        total_hours += log.hours
        categories_hours[log.category] += log.hours

        if log.category == 'engagement':
            engagement = Engagement.query.get(log.item_id)
            if engagement:
                cumulative_hours['engagements'][engagement.name] += log.hours
        elif log.category == 'proposal':
            proposal = Proposal.query.get(log.item_id)
            if proposal:
                cumulative_hours['proposals'][proposal.name] += log.hours
        elif log.category == 'non_billable':
            non_billable = NonBillable.query.get(log.item_id)
            if non_billable:
                cumulative_hours['non_billables'][non_billable.name] += log.hours

        if first_log_date is None or log.date < first_log_date:
            first_log_date = log.date
        if last_log_date is None or log.date > last_log_date:
            last_log_date = log.date

        year = log.date.year
        month = log.date.strftime('%B')
        week = log.date.isocalendar()[1]
        day = log.date.day
        hours_logs_by_date[year][month][week][day].append(log)

    today = datetime.now().date()

    # Calculate leave statistics
    current_date = datetime.now().date()

    # Determine the start and end of the current financial year
    if current_date.month > 6:
        start_of_financial_year = datetime(current_date.year, 7, 1).date()
        end_of_financial_year = datetime(current_date.year + 1, 6, 30).date()
    else:
        start_of_financial_year = datetime(current_date.year - 1, 7, 1).date()
        end_of_financial_year = datetime(current_date.year, 6, 30).date()

    # Calculate total months from start of financial year
    total_months = (current_date.year - start_of_financial_year.year) * 12 + current_date.month - start_of_financial_year.month

    # Calculate leave earned so far
    leave_earned = min(total_months * 1.75, 21)  # cap at 21 days per financial year

    # Calculate leave taken this financial year
    leave_taken = LeaveRecord.query.filter(
        LeaveRecord.staff_id == staff_id,
        LeaveRecord.date >= start_of_financial_year
    ).count()

    # Calculate current leave balance
    current_balance = max(0, leave_earned - leave_taken)

    # Calculate total leave available for the financial year
    total_leave_available = min(12 * 1.75, 21)  # cap at 21 days per financial year

    if leave_taken > 0:
        total_leave_available = max(0, total_leave_available - leave_taken)

    leave_stats = {
        'used_leave_days': leave_taken,
        'used_leave_hours': leave_taken * 7.5,  # Assuming 7.5 hours per day
        'current_balance_days': round(current_balance, 2),
        'current_balance_hours': round(current_balance * 7.5, 2),
        'total_available_days': round(total_leave_available, 2),
        'total_available_hours': round(total_leave_available * 7.5, 2)
    }

    # Calculate average daily hours
    days_with_logs = len(hours_logs_by_date)
    average_daily_hours = total_hours / days_with_logs if days_with_logs > 0 else 0

    # Calculate total workdays in the logged period
    if first_log_date and last_log_date:
        total_workdays = sum(1 for d in (first_log_date + timedelta(n) for n in range((last_log_date - first_log_date).days + 1))
                             if d.weekday() < 5)  # Monday to Friday
        available_hours = total_workdays * 7.5
    else:
        available_hours = 0

    # Calculate utilization rate
    utilization_rate = (total_hours / available_hours * 100) if available_hours > 0 else 0

    # Determine most active category
    most_active_category = max(categories_hours, key=categories_hours.get, default='none')

    return render_template('staff_detail.html',
                           staff=staff,
                           engagements=engagements,
                           proposals=proposals,
                           hours_logs=hours_logs,
                           cumulative_hours=cumulative_hours,
                           hours_logs_by_date=hours_logs_by_date,
                           Engagement=Engagement,
                           Proposal=Proposal,
                           NonBillable=NonBillable,
                           today=today,
                           total_hours=round(total_hours, 2),
                           average_daily_hours=round(average_daily_hours, 2),
                           utilization_rate=round(utilization_rate, 2),
                           most_active_category=most_active_category.capitalize(),
                           leave_stats=leave_stats)


from datetime import timedelta


@app.route('/view_logs/<int:staff_id>', methods=['GET', 'POST'])
def view_logs(staff_id):
    staff = Staff.query.get_or_404(staff_id)

    # Pagination variables
    page = request.args.get('page', 1, type=int)
    per_page = 50  # Number of logs to display per page

    logs = HoursLog.query.filter_by(staff_id=staff_id).order_by(HoursLog.date.desc()).paginate(page=page,
                                                                                               per_page=per_page)

    # Group logs by week (Monday to Friday)
    grouped_logs = {}
    for log in logs.items:
        # Get the Monday of the week for each log
        start_of_week = log.date - timedelta(days=log.date.weekday())  # Monday of the week
        end_of_week = start_of_week + timedelta(days=4)  # Friday of the week

        week_label = f"{start_of_week.strftime('%A %d')} - {end_of_week.strftime('%A %d %B %Y')}"

        if week_label not in grouped_logs:
            grouped_logs[week_label] = []
        grouped_logs[week_label].append(log)

    # Handle post requests
    if request.method == 'POST':
        if 'delete' in request.form:
            log_id = request.form['delete']
            log = HoursLog.query.get(log_id)
            if log and log.staff_id == staff_id:
                db.session.delete(log)
                db.session.commit()
                flash('Log entry deleted successfully.', 'success')
            else:
                flash('Invalid log entry.', 'error')

        # Delete all logs for a specific week
        if 'delete_week' in request.form:
            week_label = request.form['delete_week']
            logs_to_delete = grouped_logs.get(week_label, [])
            for log in logs_to_delete:
                db.session.delete(log)
            db.session.commit()
            flash(f'All logs for {week_label} deleted successfully.', 'success')

        return redirect(url_for('view_logs', staff_id=staff_id))

    return render_template('staff_logs.html', staff=staff, grouped_logs=grouped_logs, Proposal=Proposal,
                           Engagement=Engagement, NonBillable=NonBillable, logs=logs)


from flask import render_template

from datetime import timedelta
from calendar import monthrange


@app.route('/preview/<int:user_id>')
def preview_user(user_id):
    staff_member = Staff.query.get_or_404(user_id)

    # Get all hour logs and leave records for the staff member
    hours_logs = HoursLog.query.filter_by(staff_id=user_id).order_by(HoursLog.date).all()
    leave_records = {record.date for record in LeaveRecord.query.filter_by(staff_id=user_id).all()}

    # Organize logs by week (Monday to Friday)
    logs_by_week = {}
    month_set = set()  # To collect months with data

    # Fetch names for engagements, proposals, and non-billables
    engagements = {engagement.id: engagement.name for engagement in Engagement.query.all()}
    proposals = {proposal.id: proposal.name for proposal in Proposal.query.all()}
    non_billables = {non_billable.id: non_billable.name for non_billable in NonBillable.query.all()}

    daily_totals = {}  # To store daily totals
    weekly_totals = {}  # To store weekly totals

    for log in hours_logs:
        # Week Start and End (Monday to Friday)
        week_start = log.date - timedelta(days=log.date.weekday())  # Monday of the week
        week_end = week_start + timedelta(days=4)  # Friday of the week

        if week_start not in logs_by_week:
            logs_by_week[week_start] = {
                'end': week_end,
                'logs': {},
                'month': log.date.strftime('%B %Y')
            }
        if log.date not in logs_by_week[week_start]['logs']:
            logs_by_week[week_start]['logs'][log.date] = []

        # Determine item name based on category
        item_name = ""
        if log.category == 'engagement':
            item_name = engagements.get(log.item_id, 'Unknown Engagement')
        elif log.category == 'proposal':
            item_name = proposals.get(log.item_id, 'Unknown Proposal')
        elif log.category == 'non_billable':
            item_name = non_billables.get(log.item_id, 'Unknown Non-Billable')
        else:
            item_name = 'Unknown Category'

        logs_by_week[week_start]['logs'][log.date].append({
            'category': log.category,
            'hours': log.hours,
            'item_name': item_name
        })

        # Calculate daily totals
        if log.date not in daily_totals:
            daily_totals[log.date] = 0
        daily_totals[log.date] += log.hours

        # Calculate weekly totals
        if week_start not in weekly_totals:
            weekly_totals[week_start] = 0
        weekly_totals[week_start] += log.hours

        month_set.add(log.date.strftime('%B %Y'))

    # Mark days with leave
    for week_start, week_data in logs_by_week.items():
        for date in week_data['logs']:
            if date in leave_records:
                for log in week_data['logs'][date]:
                    log['item_name'] = 'User was on leave'

    # Paginate the results
    page = request.args.get('page', 1, type=int)
    per_page = 5
    paginated_weeks = list(logs_by_week.items())[(page - 1) * per_page: page * per_page]
    total_pages = (len(logs_by_week) + per_page - 1) // per_page

    return render_template('preview.html', staff_member=staff_member, logs_by_week=paginated_weeks,
                           month_list=sorted(month_set), current_page=page, total_pages=total_pages,
                           daily_totals=daily_totals, weekly_totals=weekly_totals)


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)
    